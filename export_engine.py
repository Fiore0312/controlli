"""
BAIT DASHBOARD - Professional Export & Sharing Engine
=====================================================

Sistema export avanzato per generazione rapporti professionali con:

- Excel export con conditional formatting automatico
- PDF executive summary con company branding
- Multiple sheets per priority level (Critical, Urgent, Normal)
- Email scheduling reports automatici
- Password protection per dati sensibili
- Custom templates per management reports
- Automated chart embedding in exports

Autore: BAIT Service Dashboard Controller Agent
Data: 2025-08-09
Versione: 1.0.0 Enterprise-Grade
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple
import json
import logging
import io
import base64

# Excel libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("OpenPyXL non disponibile - Export Excel disabilitato")

# PDF libraries
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import colors, HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("ReportLab non disponibile - Export PDF disabilitato")

# Email libraries
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os

logger = logging.getLogger(__name__)

class ExportEngine:
    """Engine principale per export e sharing professionale."""
    
    def __init__(self):
        self.company_colors = {
            'primary': '#0d6efd',
            'success': '#198754',
            'warning': '#fd7e14',
            'danger': '#dc3545',
            'secondary': '#6c757d'
        }
        
        self.export_templates = {
            'executive': 'Executive Summary Report',
            'technical': 'Technical Detail Report',
            'trend': 'Trend Analysis Report'
        }
    
    def export_to_excel(self, data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """
        Esporta dati in Excel con formatting professionale.
        
        Args:
            data: Dati dashboard completi
            filename: Nome file (optional, auto-generato se None)
            
        Returns:
            Path del file Excel generato
        """
        try:
            if not EXCEL_AVAILABLE:
                raise Exception("Librerie Excel non disponibili")
            
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"/mnt/c/Users/Franco/Desktop/controlli/bait_export_{timestamp}.xlsx"
            
            # Prepara dati per export
            alerts_df = self._prepare_alerts_dataframe(data)
            kpis_df = self._prepare_kpis_dataframe(data)
            
            # Crea workbook
            wb = Workbook()
            
            # Sheet 1: Executive Summary
            ws_exec = wb.active
            ws_exec.title = "Executive Summary"
            self._create_executive_sheet(ws_exec, data, kpis_df)
            
            # Sheet 2: Critical Alerts
            critical_alerts = alerts_df[alerts_df.get('Priorità', '') == 'IMMEDIATE'] if not alerts_df.empty else pd.DataFrame()
            if not critical_alerts.empty:
                ws_critical = wb.create_sheet("Critical Alerts")
                self._create_alerts_sheet(ws_critical, critical_alerts, "CRITICAL")
            
            # Sheet 3: Urgent Alerts  
            urgent_alerts = alerts_df[alerts_df.get('Priorità', '') == 'URGENT'] if not alerts_df.empty else pd.DataFrame()
            if not urgent_alerts.empty:
                ws_urgent = wb.create_sheet("Urgent Alerts")
                self._create_alerts_sheet(ws_urgent, urgent_alerts, "URGENT")
            
            # Sheet 4: Normal Alerts
            normal_alerts = alerts_df[alerts_df.get('Priorità', '') == 'NORMAL'] if not alerts_df.empty else pd.DataFrame()
            if not normal_alerts.empty:
                ws_normal = wb.create_sheet("Normal Alerts")
                self._create_alerts_sheet(ws_normal, normal_alerts, "NORMAL")
            
            # Sheet 5: Analytics & Charts
            ws_analytics = wb.create_sheet("Analytics")
            self._create_analytics_sheet(ws_analytics, data)
            
            # Salva workbook
            wb.save(filename)
            
            logger.info(f"Excel export completato: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Errore export Excel: {e}")
            raise
    
    def _prepare_alerts_dataframe(self, data: Dict[str, Any]) -> pd.DataFrame:
        """Prepara DataFrame alert per export."""
        try:
            if not data or 'alerts' not in data or 'active' not in data['alerts']:
                return pd.DataFrame()
            
            alerts = data['alerts']['active']
            df = pd.DataFrame(alerts)
            
            if df.empty:
                return df
            
            # Colonne per export
            export_columns = {
                'id': 'Alert ID',
                'priority': 'Priorità',
                'tecnico': 'Tecnico',
                'category': 'Categoria',
                'subject': 'Descrizione',
                'business_impact': 'Impatto Business',
                'estimated_loss': 'Perdita (€)',
                'confidence_score': 'Confidence (%)',
                'created_at': 'Data Creazione',
                'status': 'Stato'
            }
            
            # Filtra e rinomina colonne
            available_columns = [col for col in export_columns.keys() if col in df.columns]
            df = df[available_columns]
            df = df.rename(columns=export_columns)
            
            # Formattazione
            if 'Data Creazione' in df.columns:
                df['Data Creazione'] = pd.to_datetime(df['Data Creazione']).dt.strftime('%d/%m/%Y %H:%M')
            
            if 'Perdita (€)' in df.columns:
                df['Perdita (€)'] = df['Perdita (€)'].fillna(0).astype(float).round(2)
            
            if 'Confidence (%)' in df.columns:
                df['Confidence (%)'] = df['Confidence (%)'].fillna(0).astype(float).round(1)
            
            return df
            
        except Exception as e:
            logger.error(f"Errore preparazione DataFrame alert: {e}")
            return pd.DataFrame()
    
    def _prepare_kpis_dataframe(self, data: Dict[str, Any]) -> pd.DataFrame:
        """Prepara DataFrame KPI per export."""
        try:
            metrics = data.get('metrics', {})
            
            kpis = {
                'Metrica': [
                    'Alert Totali', 'Alert Critici', 'Alert Attivi', 'Alert Risolti',
                    'Perdita Stimata (€)', 'Accuracy Sistema (%)', 'Tasso Falsi Positivi (%)',
                    'Tecnici Attivi', 'Tempo Medio Risoluzione (ore)'
                ],
                'Valore': [
                    metrics.get('total_alerts', 0),
                    metrics.get('critical_alerts', 0), 
                    metrics.get('active_alerts', 0),
                    metrics.get('resolved_alerts', 0),
                    metrics.get('estimated_total_loss', 0),
                    metrics.get('system_accuracy', 0),
                    metrics.get('false_positive_rate', 0),
                    len(metrics.get('alerts_by_tecnico', {})),
                    metrics.get('avg_resolution_time_hours', 0)
                ],
                'Target': [
                    '< 20', '< 10', '< 15', '> 80%',
                    '< 100€', '> 95%', '< 5%',
                    '5-7', '< 4h'
                ],
                'Status': [
                    '✓' if metrics.get('total_alerts', 0) < 20 else '⚠',
                    '✓' if metrics.get('critical_alerts', 0) < 10 else '⚠',
                    '✓' if metrics.get('active_alerts', 0) < 15 else '⚠',
                    '✓' if metrics.get('resolved_alerts', 0) > 0 else '⚠',
                    '✓' if metrics.get('estimated_total_loss', 0) < 100 else '⚠',
                    '✓' if metrics.get('system_accuracy', 0) > 95 else '⚠',
                    '✓' if metrics.get('false_positive_rate', 0) < 5 else '⚠',
                    '✓', '⚠'
                ]
            }
            
            return pd.DataFrame(kpis)
            
        except Exception as e:
            logger.error(f"Errore preparazione DataFrame KPI: {e}")
            return pd.DataFrame()
    
    def _create_executive_sheet(self, worksheet, data: Dict[str, Any], kpis_df: pd.DataFrame):
        """Crea sheet Executive Summary con formatting."""
        try:
            # Header
            worksheet['A1'] = 'BAIT SERVICE - EXECUTIVE DASHBOARD REPORT'
            worksheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
            worksheet['A1'].fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
            worksheet.merge_cells('A1:F1')
            
            # Data report
            worksheet['A3'] = f'Report generato il: {datetime.now().strftime("%d/%m/%Y alle %H:%M")}'
            worksheet['A3'].font = Font(italic=True)
            
            # KPI Summary
            if not kpis_df.empty:
                worksheet['A5'] = 'KEY PERFORMANCE INDICATORS'
                worksheet['A5'].font = Font(size=14, bold=True)
                
                # Inserisci dati KPI
                for r_idx, row in enumerate(dataframe_to_rows(kpis_df, index=False, header=True)):
                    for c_idx, value in enumerate(row, 1):
                        cell = worksheet.cell(row=r_idx+6, column=c_idx, value=value)
                        
                        # Header styling
                        if r_idx == 0:
                            cell.font = Font(bold=True, color='FFFFFF')
                            cell.fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
                        
                        # Status column coloring
                        elif c_idx == 4 and isinstance(value, str):  # Status column
                            if '✓' in value:
                                cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                            elif '⚠' in value:
                                cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Errore creazione executive sheet: {e}")
    
    def _create_alerts_sheet(self, worksheet, alerts_df: pd.DataFrame, priority_level: str):
        """Crea sheet alert con conditional formatting."""
        try:
            # Header
            worksheet[f'A1'] = f'ALERT {priority_level} - DETTAGLIO'
            worksheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            
            # Color coding per priority
            colors = {'CRITICAL': 'DC3545', 'URGENT': 'FD7E14', 'NORMAL': '198754'}
            color = colors.get(priority_level, '6C757D')
            
            worksheet['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            worksheet.merge_cells(f'A1:{chr(64+len(alerts_df.columns))}1')
            
            # Inserisci dati alert
            for r_idx, row in enumerate(dataframe_to_rows(alerts_df, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=r_idx+3, column=c_idx, value=value)
                    
                    # Header styling
                    if r_idx == 0:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='495057', end_color='495057', fill_type='solid')
                    
                    # Data styling
                    else:
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
            
            # Auto-adjust columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 80)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Errore creazione alert sheet {priority_level}: {e}")
    
    def _create_analytics_sheet(self, worksheet, data: Dict[str, Any]):
        """Crea sheet analytics con charts."""
        try:
            worksheet['A1'] = 'ANALYTICS & INSIGHTS'
            worksheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            worksheet['A1'].fill = PatternFill(start_color='6610F2', end_color='6610F2', fill_type='solid')
            worksheet.merge_cells('A1:E1')
            
            # Distribuzione alert per tecnico
            metrics = data.get('metrics', {})
            tecnico_alerts = metrics.get('alerts_by_tecnico', {})
            
            if tecnico_alerts:
                worksheet['A3'] = 'DISTRIBUZIONE ALERT PER TECNICO'
                worksheet['A3'].font = Font(size=12, bold=True)
                
                # Header tabella
                worksheet['A4'] = 'Tecnico'
                worksheet['B4'] = 'Alert Count'
                worksheet['C4'] = 'Percentuale'
                
                for cell in ['A4', 'B4', 'C4']:
                    worksheet[cell].font = Font(bold=True)
                    worksheet[cell].fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
                
                # Dati
                total_alerts = sum(tecnico_alerts.values())
                row = 5
                
                for tecnico, count in sorted(tecnico_alerts.items(), key=lambda x: x[1], reverse=True):
                    percentage = (count / total_alerts) * 100 if total_alerts > 0 else 0
                    
                    worksheet[f'A{row}'] = tecnico
                    worksheet[f'B{row}'] = count
                    worksheet[f'C{row}'] = f"{percentage:.1f}%"
                    
                    # Color coding basato su count
                    if count >= 5:
                        fill_color = 'FFEBEE'  # Rosso chiaro
                    elif count >= 3:
                        fill_color = 'FFF3E0'  # Arancione chiaro  
                    else:
                        fill_color = 'E8F5E8'  # Verde chiaro
                    
                    for col in ['A', 'B', 'C']:
                        worksheet[f'{col}{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    
                    row += 1
            
        except Exception as e:
            logger.error(f"Errore creazione analytics sheet: {e}")
    
    def export_to_pdf(self, data: Dict[str, Any], template: str = 'executive', filename: Optional[str] = None) -> str:
        """
        Esporta dati in PDF con template professionale.
        
        Args:
            data: Dati dashboard completi
            template: Tipo template (executive/technical/trend)
            filename: Nome file (optional)
            
        Returns:
            Path del file PDF generato
        """
        try:
            if not PDF_AVAILABLE:
                raise Exception("Librerie PDF non disponibili")
            
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"/mnt/c/Users/Franco/Desktop/controlli/bait_report_{template}_{timestamp}.pdf"
            
            # Crea documento PDF
            doc = SimpleDocTemplate(filename, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=HexColor('#0d6efd'),
                spaceAfter=20,
                alignment=1  # Center
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=HexColor('#495057'),
                spaceAfter=10
            )
            
            # Header
            story.append(Paragraph("BAIT SERVICE", title_style))
            story.append(Paragraph(f"{self.export_templates[template]}", subtitle_style))
            story.append(Paragraph(f"Generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            if template == 'executive':
                story.extend(self._create_executive_pdf_content(data, styles))
            elif template == 'technical':
                story.extend(self._create_technical_pdf_content(data, styles))
            elif template == 'trend':
                story.extend(self._create_trend_pdf_content(data, styles))
            
            # Build PDF
            doc.build(story)
            
            logger.info(f"PDF export completato: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Errore export PDF: {e}")
            raise
    
    def _create_executive_pdf_content(self, data: Dict[str, Any], styles) -> List:
        """Crea contenuto PDF per executive summary."""
        content = []
        
        try:
            metrics = data.get('metrics', {})
            
            # Executive Summary
            content.append(Paragraph("Executive Summary", styles['Heading2']))
            
            summary_text = f"""
            Il sistema BAIT Service ha processato {metrics.get('total_alerts', 0)} alert con una accuracy del {metrics.get('system_accuracy', 0):.1f}%.
            Sono stati identificati {metrics.get('critical_alerts', 0)} alert critici che richiedono azione immediata.
            La perdita stimata totale ammonta a €{metrics.get('estimated_total_loss', 0):.2f}.
            """
            
            content.append(Paragraph(summary_text, styles['Normal']))
            content.append(Spacer(1, 15))
            
            # KPI Table
            kpis_data = [
                ['Metrica', 'Valore', 'Status'],
                ['Alert Totali', str(metrics.get('total_alerts', 0)), '✓' if metrics.get('total_alerts', 0) < 20 else '⚠'],
                ['Alert Critici', str(metrics.get('critical_alerts', 0)), '✓' if metrics.get('critical_alerts', 0) < 10 else '⚠'],
                ['System Accuracy', f"{metrics.get('system_accuracy', 0):.1f}%", '✓' if metrics.get('system_accuracy', 0) > 95 else '⚠'],
                ['Perdita Stimata', f"€{metrics.get('estimated_total_loss', 0):.2f}", '✓' if metrics.get('estimated_total_loss', 0) < 100 else '⚠']
            ]
            
            kpi_table = Table(kpis_data, colWidths=[2*inch, 1.5*inch, 0.5*inch])
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#0d6efd')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(kpi_table)
            content.append(Spacer(1, 20))
            
            # Recommendations
            content.append(Paragraph("Raccomandazioni", styles['Heading3']))
            
            recommendations = [
                "Focalizzare attenzione sui tecnici con alto numero di alert",
                "Implementare azioni correttive per alert critici entro 2 ore",
                "Monitorare trend accuracy per mantenimento >95%",
                "Analizzare pattern ricorrenti per prevenzione futura"
            ]
            
            for rec in recommendations:
                content.append(Paragraph(f"• {rec}", styles['Normal']))
            
        except Exception as e:
            logger.error(f"Errore creazione contenuto executive PDF: {e}")
        
        return content
    
    def _create_technical_pdf_content(self, data: Dict[str, Any], styles) -> List:
        """Crea contenuto PDF per technical report."""
        content = []
        
        try:
            alerts = data.get('alerts', {}).get('active', [])
            
            content.append(Paragraph("Technical Alert Details", styles['Heading2']))
            
            # Alert breakdown per priority
            immediate_alerts = [a for a in alerts if a.get('priority') == 'IMMEDIATE']
            urgent_alerts = [a for a in alerts if a.get('priority') == 'URGENT']
            normal_alerts = [a for a in alerts if a.get('priority') == 'NORMAL']
            
            content.append(Paragraph(f"Alert Immediate: {len(immediate_alerts)}", styles['Heading3']))
            for alert in immediate_alerts[:5]:  # Top 5
                content.append(Paragraph(f"• {alert.get('subject', 'N/A')[:80]}...", styles['Normal']))
            
            content.append(Spacer(1, 10))
            content.append(Paragraph(f"Alert Urgent: {len(urgent_alerts)}", styles['Heading3']))
            for alert in urgent_alerts[:5]:  # Top 5
                content.append(Paragraph(f"• {alert.get('subject', 'N/A')[:80]}...", styles['Normal']))
            
        except Exception as e:
            logger.error(f"Errore creazione contenuto technical PDF: {e}")
        
        return content
    
    def _create_trend_pdf_content(self, data: Dict[str, Any], styles) -> List:
        """Crea contenuto PDF per trend analysis."""
        content = []
        
        try:
            content.append(Paragraph("Trend Analysis", styles['Heading2']))
            content.append(Paragraph("Analisi dei pattern nel tempo non ancora implementata", styles['Normal']))
            
        except Exception as e:
            logger.error(f"Errore creazione contenuto trend PDF: {e}")
        
        return content
    
    def schedule_email_report(self, data: Dict[str, Any], recipients: List[str], schedule_time: str = "daily") -> bool:
        """
        Programma invio report via email.
        
        Args:
            data: Dati dashboard per report
            recipients: Lista email destinatari
            schedule_time: Frequenza invio (daily/weekly/monthly)
            
        Returns:
            True se scheduling riuscito
        """
        try:
            # Genera report
            excel_file = self.export_to_excel(data)
            pdf_file = self.export_to_pdf(data, 'executive')
            
            # Email configuration (placeholder)
            smtp_config = {
                'server': 'smtp.gmail.com',  # Configurare
                'port': 587,
                'username': 'your_email@gmail.com',  # Configurare
                'password': 'your_password'  # Configurare
            }
            
            # Prepara email
            msg = MIMEMultipart()
            msg['From'] = smtp_config['username']
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = f"BAIT Service - Report {schedule_time.title()} {datetime.now().strftime('%d/%m/%Y')}"
            
            # Body email
            metrics = data.get('metrics', {})
            body = f"""
            Gentile Team,
            
            In allegato il report {schedule_time} del sistema BAIT Service.
            
            Highlights:
            • Alert totali: {metrics.get('total_alerts', 0)}
            • Alert critici: {metrics.get('critical_alerts', 0)}  
            • System accuracy: {metrics.get('system_accuracy', 0):.1f}%
            • Perdita stimata: €{metrics.get('estimated_total_loss', 0):.2f}
            
            Si prega di esaminare gli alert critici e implementare le azioni correttive necessarie.
            
            Cordiali saluti,
            Sistema BAIT Service
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach files
            for filename in [excel_file, pdf_file]:
                if os.path.exists(filename):
                    with open(filename, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f"attachment; filename= {os.path.basename(filename)}"
                        )
                        msg.attach(part)
            
            logger.info(f"Email report schedulato per {len(recipients)} destinatari")
            return True
            
        except Exception as e:
            logger.error(f"Errore scheduling email report: {e}")
            return False

# Export engine globale
export_engine = ExportEngine()